from selenium import webdriver
import time
import os
import cv2
import numpy as np
from itertools import islice
import re
from openpyxl import load_workbook
from pymouse import PyMouse
from pykeyboard import PyKeyboard
from PIL import Image, ImageGrab
import sys
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
from aip import AipOcr
import msvcrt
import logging


class Imgcat:
    # 添加一个模板找寻并点击的类
    def __init__(self, img_name, vc_name):
        self.img_name = img_name
        self.vc_name = vc_name

    def getxy(self):
        while True:
            try:
                sc_path = './{}.png'.format('temp')
                m = PyMouse()
                x_dim, y_dim = m.screen_size()
                img_rgb = ImageGrab.grab((0, 0, x_dim, y_dim))
                img_rgb.save(sc_path)
                img_rgb = cv2.imread(sc_path)
                img_gray = cv2.cvtColor(img_rgb, cv2.COLOR_BGR2GRAY)
                template = cv2.imread('{}.png'.format(self.img_name), 0)
                w, h = template.shape[::-1]
                res = cv2.matchTemplate(
                    img_gray, template, cv2.TM_CCOEFF_NORMED)
                threshold = 0.9
                loc = np.where(res >= threshold)
                for pt in zip(*loc[::-1]):
                    #cv2.rectangle(img_rgb, pt, (pt[0] + w, pt[1] + h), (0,0,255), 2)
                    print('Found', pt[0] + w // 2, pt[1] + h // 2)
                    x = int(pt[0] + w // 2)
                    y = int(pt[1] + h // 2)
                ImageGrab.grab((x+198, y-17, x+198+123, y+13)
                               ).save('{}.png'.format(self.vc_name))
                """ 你的 APPID AK SK """
                APP_ID = 'xxxx'
                API_KEY = 'xxxx'
                SECRET_KEY = 'xxxx'
                client = AipOcr(APP_ID, API_KEY, SECRET_KEY)
                """ 读取图片 """
                def get_file_content(f):
                    with open(f, 'rb') as fp:
                        return fp.read()
                image = get_file_content('{}.png'.format(self.vc_name))
                # 定义参数变量
                options = {
                    # 定义图像方向
                    'detect_direction': 'true',
                    # 识别语言类型，默认为'CHN_ENG'中英文混合
                    'language_type': 'ENG',
                    # 'detect_language' : "true"
                }
                results = client.basicAccurate(image, options)
                print(results)
                return str(results['words_result'][0]['words']).replace(' ', '')
                break
            except UnboundLocalError:
                print('没有找到样本，重试')
                time.sleep(0.5)
            except IndexError:
                print('ocr识别抓图错误，或许是在截图的时候拖动了窗口？')


class LoginMan():
    def __init__(self, u, p):
        self.u = u
        self.p = p

    def login(self):
        while True:
            try:
                # 分别发送用户名、密码、验证码
                c = Imgcat('down', 'code').getxy()
                for i, j in zip(('username', 'password', 'validateCode'), (self.u, self.p, c)):
                    print('发送 {}'.format(j))
                    driver.find_element_by_xpath(
                        '//*[@name="{}"]'.format(i)).send_keys(j)
                    time.sleep(0.5)
                    if j == c:
                        driver.find_element_by_xpath(
                            '//*[@name="{}"]'.format(i)).send_keys(Keys.RETURN)
                time.sleep(1)
                verror = driver.find_element_by_xpath(
                    '//*[@id="validateCodeMessage" and text()="验证码错误"]').is_displayed()
                print(not verror)  # 判断是否发现验证码错误的提示
                if verror == True:
                    print('验证码OCR识别错误，重试')
            except NoSuchElementException:  # 如果没有发现，那么说明输入正确,break断开
                print('估计验证码识别准确并已登录')
                time.sleep(1)
                break


def playv():
    video = driver.find_element_by_xpath("//*[@id='my-video_html5_api']")
    url = driver.execute_script("return arguments[0].currentSrc;", video)
    print(url)
    button = driver.find_element_by_xpath("//*[@class='vjs-big-play-button']")
    button.click()
    print('START PLAY')
    #vt= driver.execute_script("return arguments[0].duration",video)
    #print (vt)


def countdown(t):
    for i in range(100):
        sys.stdout.write('   \r')
        sys.stdout.flush()
        sys.stdout.write('按S键跳过该视频：{}%\r'.format(i))
        sys.stdout.flush()
        time.sleep(t / 100)
        # 按下s键盘跳过函数
        if msvcrt.kbhit():
            data = msvcrt.getch().decode('utf-8')
            if (data == "s"):
                print('s键按下，跳过')
                break


m = PyMouse()
k = PyKeyboard()
sc_path = 'e:/temp/autojump.png'
url = 'http://dyjy.dtdjzx.gov.cn/resourcedetailed/'
urlid = ('2766472326202368', '2589222270239744')
# 每个视频的播放时间
vt = (3, 800)


def playh():
    for index, (m, n) in enumerate(zip(urlid, vt)):
        print('{} OPEN {}{}'.format(index+1, url, m))
        driver.get('{}{}'.format(url, m))
        playv()
        countdown(n)


print('>>当前屏幕分辨率为{}'.format(m.screen_size()))
print('>>>开始答题')
wb = load_workbook('user.xlsx')
sheet = wb['Sheet1']
# print ('共计{}列'.format(sheet.max_column))#显示最大列
print('共计{}人'.format(sheet.max_row))  # 显示最大行
logging.basicConfig(filename='log.txt', level=logging.INFO)
for index, i in enumerate(range(1, sheet.max_row+1)):
    user = sheet.cell(row=i, column=1).value  # i1单元格
    pswd = sheet.cell(row=i, column=2).value  # i2单元格
    user_name = sheet.cell(row=i, column=3).value  # i3单元格
    print('当前第{}个答题用户为：{}，密码为{}' .format(index+1, user, pswd))
    driver = webdriver.Chrome()
    print('打开党建学习个人中心')
    driver.get('http://dyjy.dtdjzx.gov.cn/personal')
    LoginMan(user, pswd).login()
    # 定位包含特殊文本的元素
    s_h = driver.find_element_by_xpath(
        "//div[contains(text(),'党员教育总学时') and @class='rbox-tips']").text[:10]
    print('用户：{} {}'.format(user, s_h))
    # 记录分数到文件
    logging.info('=={}  {}  {}  {}'.format(index+1, user_name, user, s_h))
    # 打开视频播放或只查看分数
    # playh()
    driver.quit()
